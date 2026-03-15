# fill_saisie_interet.py
#
# Version 0.1.1 (2025/12/08)
# Bugfix: correctly write into merged cells by targeting the top-left cell of each merged range.

from pathlib import Path
from typing import Dict, Union

import csv
import openpyxl
from openpyxl.cell.cell import MergedCell

CellValue = Union[str, int, float]


def get_dummy_data() -> Dict[int, CellValue]:
    """
    Dummy values for template testing.
    Keys are Excel row numbers for column C.
    NOTE: C31 (Ionscan #) and C35 (Nom du chien detecteur) are intentionally omitted.
    """
    return {
        7: "2025-03-11",
        8: "12345",
        9: "Recommande",
        10: "ABXXXYYYZZZCO",
        12: "XXXXXXXXXX",
        13: "1234g",
        14: "12345$",
        15: "2 small packets and large wrapped in plastic, concealed in regular size (4 X 9.5) envelope.",
        17: "John Doe",
        18: "12345 One str, Ottawa, Canada K1A 1A1",
        19: "N/A",
        21: "Joanne Doah",
        22: "1234 Large str, NY, 05555, USA",
        23: "N/A",
        25: "N/A",
        26: 3,
        27: 22,
        29: "Aucun",
        30: "Non",
        32: "Oui",
        33: "XXXXXXXX",
        34: "Non",
        37: "folder_name",
        38: "RAS",
    }


def load_data_from_csv(csv_path: Union[str, Path]) -> Dict[int, CellValue]:
    """
    Read data from a CSV file with at least two columns: row,value

    Example:
      row,value
      7,2025-03-11
      8,12345
    """
    csv_path = Path(csv_path)
    data: Dict[int, CellValue] = {}

    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = [x.strip().lower() for x in (reader.fieldnames or [])]
        if ("row" not in fieldnames) or ("value" not in fieldnames):
            raise ValueError("CSV must have columns 'row' and 'value'.")

        for line in reader:
            row_str = (line.get("row") or "").strip()
            value_str = line.get("value", "")
            if not row_str:
                continue

            row = int(row_str)

            v: CellValue
            try:
                v = int(value_str)
            except ValueError:
                try:
                    v = float(value_str)
                except ValueError:
                    v = value_str
            data[row] = v

    return data


def _set_value_in_col_c_merged_safe(ws, row: int, value: CellValue) -> None:
    """
    Set a value in column C at the given row, handling merged cells safely.
    If C{row} is part of a merged range, write to the top-left cell of that range.
    """
    col = 3  # C
    cell = ws.cell(row=row, column=col)

    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    coord = cell.coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left.value = value
            return


def fill_saisie_interet(
    template_xlsx: Union[str, Path],
    output_xlsx: Union[str, Path],
    data_by_row: Dict[int, CellValue],
    sheet_name: str = None,
) -> None:
    """
    Fill the SAISIE D'INTERET Excel template.

    - Fills values in column C for rows present in data_by_row.
    - C31 (Ionscan #) and C35 (Nom du chien detecteur) are never overwritten.
    """
    template_xlsx = Path(template_xlsx)
    output_xlsx = Path(output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_xlsx)

    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {template_xlsx}")
        ws = wb[sheet_name]

    skip_rows = {31, 35}

    for row, value in data_by_row.items():
        if row in skip_rows:
            continue
        _set_value_in_col_c_merged_safe(ws, row=row, value=value)

    wb.save(output_xlsx)


if __name__ == "__main__":
    template = "Saisie d'interet - TEMPLATE.xlsx"
    csv_file = "saisie_interet_values.csv"
    output_csv = "Saisie d'interet_filled_from_csv.xlsx"
    csv_data = load_data_from_csv(csv_file)
    fill_saisie_interet(
        template_xlsx=template,
        output_xlsx=output_csv,
        data_by_row=csv_data,
        sheet_name=None,
    )
